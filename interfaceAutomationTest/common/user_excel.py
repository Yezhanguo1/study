import os
import sys
sys.path.append('..')
sys.path.append(os.path.dirname(sys.path[0]))
from openpyxl import Workbook, load_workbook


class Excel:
    def read_case_excel(self, file_path, Sheel_name):
        """
        默认读取跳过第一行表头的数据，按行读取
        :rtype: str
        :param file_path:Excel文件路径
        :param Sheel_name:工作薄sheel名称
        :return:用例参数列表
        """
        wk = load_workbook(file_path)
        sd = wk.get_sheet_by_name(Sheel_name)
        case_param_list = []
        for lin_date in sd.iter_rows(min_row=2,min_col=2,values_only=True):
            case_param_list.append(lin_date)
        return case_param_list


# if __name__ == '__main__':
#     a=Excel.read_case_excel(Excel(),file_path="../case/test_case_date.xlsx",Sheel_name="denglu")

